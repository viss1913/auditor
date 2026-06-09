# -*- coding: utf-8 -*-
"""Создаёт Бланк_спецификации_парсинга_ОС.xlsx в корне проекта."""
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    import sys

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "Бланк_спецификации_парсинга_ОС.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
BOLD = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def set_header_row(ws, row, values, widths=None):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.alignment = WRAP
    if widths:
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w


wb = openpyxl.Workbook()

# --- Лист: Инструкция ---
ws0 = wb.active
ws0.title = "Инструкция"
lines = [
    "Бланк спецификации парсинга ОС / капвложений (для аудитора)",
    "",
    "Заполните листы «Целевая_01» и/или «Целевая_08» по образцу.",
    "Приложите к ТЗ: 1) исходную выгрузку 1С; 2) этот бланк (заполненный); 3) пример 2–3 строк вручную.",
    "",
    "Правило для разработки = исходник + целевые колонки + пояснения на листе «Вопросы».",
    "Минимум этапа 1: плоская таблица (одна строка = один объект × период × набор метрик).",
    "",
    "Единицы: укажите рубли или тысячи. Подитоги (Итого, группы) — не выводить, если не оговорено.",
]
for i, line in enumerate(lines, 1):
    ws0.cell(row=i, column=1, value=line)
ws0.column_dimensions["A"].width = 100

# --- Целевая_01 ---
ws1 = wb.create_sheet("Целевая_01")
h01 = [
    "Юрлицо",
    "Год отчёта",
    "Группа учёта ОС",
    "Подразделение (ОП/РТК)",
    "Наименование ОС",
    "Инв. №",
    "Дата принятия к учёту",
    "Стоимость на начало",
    "Амортизация на начало",
    "Остаточная на начало",
    "Увеличение стоимости",
    "Начисление амортизации",
    "Уменьшение стоимости",
    "Списание амортизации",
    "Стоимость на конец",
    "Амортизация на конец",
    "Остаточная на конец",
    "Комментарий аудитора",
]
set_header_row(ws1, 1, h01, [12, 10, 22, 18, 45, 14, 14] + [16] * 10 + [30])
example01 = [
    "ОАО",
    2024,
    "Здания",
    "ОП КЦ",
    "Модульное здание 3м*10,6м, 80-000722, 23.04.2012",
    "80-000722",
    "23.04.2012",
    0,
    0,
    0,
    1100000,
    103529.44,
    None,
    None,
    1100000,
    103529.44,
    996470.56,
    "Пример строки — заполните свои",
]
for col, val in enumerate(example01, 1):
    ws1.cell(row=2, column=col, value=val)

# --- Целевая_08 ---
ws2 = wb.create_sheet("Целевая_08")
h08 = [
    "Юрлицо",
    "Счёт",
    "Подразделение",
    "Объект / описание",
    "Период (как в выгрузке)",
    "Год",
    "Сальдо Дт начало",
    "Сальдо Кт начало",
    "Оборот Дт",
    "Оборот Кт",
    "Сальдо Дт конец",
    "Сальдо Кт конец",
    "Комментарий",
]
set_header_row(ws2, 1, h08, [12, 10, 18, 50, 22, 8] + [14] * 6 + [25])
example08 = [
    "ОАО",
    "08",
    "ОП Волгоград",
    "Дооборудование Сервер с ИБП инв. №80-000662",
    "Обороты за 2023",
    2023,
    None,
    None,
    19300,
    19300,
    None,
    None,
    "Пример",
]
for col, val in enumerate(example08, 1):
    ws2.cell(row=2, column=col, value=val)

# --- Вопросы ---
ws3 = wb.create_sheet("Вопросы_аудитору")
questions = [
    ("№", "Вопрос", "Ответ аудитора"),
    (1, "Одна строка результата = что? (объект за год / за период / иное)", ""),
    (2, "Единицы измерения в отчёте (руб / тыс. руб)", ""),
    (3, "Какие строки исходника НЕ выводить (подитоги, группы, пустые)?", ""),
    (4, "Фиксированный список групп ОС или определяется из файла?", ""),
    (5, "Нужно ли разбивать ячейку «название, инв.№, дата» на отдельные поля?", ""),
    (6, "Для сч. 08: какие колонки ОСВ использовать (остаток, оборот Дт/Кт)?", ""),
    (7, "Имя листа / шаблон 1С (если известно)", ""),
]
for r, row in enumerate(questions, 1):
    for c, val in enumerate(row, 1):
        cell = ws3.cell(row=r, column=c, value=val)
        if r == 1:
            cell.font = BOLD
            cell.fill = HEADER_FILL
ws3.column_dimensions["A"].width = 5
ws3.column_dimensions["B"].width = 55
ws3.column_dimensions["C"].width = 45

# --- Маппинг колонок ---
ws4 = wb.create_sheet("Маппинг_исходник")
set_header_row(
    ws4,
    1,
    [
        "Вариант",
        "Поле в целевой таблице",
        "Откуда в 1С (название блока/колонки)",
        "Номер колонки в Excel (если известен)",
        "Преобразование",
    ],
    [14, 28, 40, 12, 30],
)
mapping_rows = [
    ("01", "period_year", "Заголовок «за 20XX г.»", "текст шапки", "regex год"),
    ("01", "group_name", "Группа учёта ОС", "col B, уровень 1", "иерархия"),
    ("01", "subdivision", "ОП / РТК", "col B", "иерархия"),
    ("01", "asset_name", "Основное средство…", "col B, листовая строка", ""),
    ("01", "amort_charge", "Начисление амортизации (износа)", "col G", "число"),
    ("08", "turnover_dt", "Обороты за период — Дебет", "col D", "число"),
    ("08", "object_name", "Наименование объекта", "col A", "строка перед «Обороты за…»"),
]
for i, row in enumerate(mapping_rows, 2):
    for c, val in enumerate(row, 1):
        ws4.cell(row=i, column=c, value=val)

wb.save(OUT)
print(f"Создан: {OUT}")

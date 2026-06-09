# -*- coding: utf-8 -*-
"""Читает .xlsx/.xlsm через openpyxl: бары, цвета, мержи, скрытые строки → JSON в stdout."""
from __future__ import annotations

import json
import re
import sys
import os
from pathlib import Path

STYLE_SCAN_LIMIT = int(os.environ.get("EXCEL_PROBE_STYLE_SCAN_ROWS", "400"))

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print(json.dumps({"ok": False, "error": "openpyxl not installed"}, ensure_ascii=False))
    sys.exit(1)

GRAY_FILLS = {
    "FFD9D9D9",
    "FFD9E1F2",
    "FFF2F2F2",
    "FFE7E6E6",
    "FFBFBFBF",
    "FFCCCCCC",
    "FFC0C0C0",
    "FFE0E0E0",
    "FFF0F0F0",
    "FFD0CECE",
    "FFDBDBDB",
}

HEADER_FILL_HINTS = {
    "FFD9E1F2",
    "FFB4C6E7",
    "FF4472C4",
    "FF8EA9DB",
}

# Типичные заливки иерархии в выгрузках 1С (зелёные оттенки)
HIERARCHY_FILLS = {
    "FFD6E5CB",
    "FFE4F0DD",
    "FFF0F6EF",
    "FFC6E0B4",
    "FFB8D4A8",
}


def _norm_rgb(raw: str | None) -> str | None:
    if not raw:
        return None
    s = str(raw).strip().upper()
    if len(s) == 6:
        s = "FF" + s
    if len(s) == 8:
        return s
    return None


def cell_fill_rgb(cell) -> str | None:
    fill = getattr(cell, "fill", None)
    if not fill or getattr(fill, "fill_type", None) != "solid":
        return None
    fg = getattr(fill, "fgColor", None)
    if not fg:
        return None
    rgb = getattr(fg, "rgb", None)
    if rgb:
        return _norm_rgb(rgb)
    return None


def cell_is_bold(cell) -> bool:
    font = getattr(cell, "font", None)
    return bool(getattr(font, "bold", False)) if font else False


def pick_sheet(wb, explicit: str | None):
    names = wb.sheetnames
    if explicit and explicit in names:
        return explicit
    for s in names:
        if re.search(r"исходн.*осв", s, re.I):
            return s
    for s in names:
        if re.search(r"исходн", s, re.I) and not re.search(r"кс", s, re.I):
            return s
    for s in names:
        if re.search(r"осв", s, re.I) and not re.search(r"кс", s, re.I):
            return s
    for s in names:
        if re.search(r"исходн", s, re.I):
            return s
    return names[0] if names else None


def row_label(ws, row_idx: int, name_col: int = 1) -> str:
    val = ws.cell(row=row_idx, column=name_col).value
    return str(val).strip() if val is not None else ""


def row_has_numbers(ws, row_idx: int, from_col: int = 2, to_col: int | None = None) -> bool:
    max_col = to_col or min(ws.max_column or 1, 20)
    for col in range(from_col, max_col + 1):
        val = ws.cell(row=row_idx, column=col).value
        if isinstance(val, (int, float)) and val != 0:
            return True
        if val is None:
            continue
        s = str(val).replace(" ", "").replace("\u00a0", "").replace(",", ".")
        if re.match(r"^-?\d", s):
            return True
    return False


def build_style_hints(ws, row_count: int, scan_limit: int | None = None) -> dict:
    likely_subtotal_rows: list[int] = []
    likely_header_rows: list[int] = []
    gray_fill_rows: list[int] = []
    hierarchy_fill_rows: list[int] = []
    hidden_rows: list[int] = []

    limit = scan_limit if scan_limit is not None else min(row_count, STYLE_SCAN_LIMIT)

    for i in range(1, row_count + 1):
        rd = ws.row_dimensions.get(i)
        hidden = bool(getattr(rd, "hidden", False)) if rd else False
        if hidden:
            hidden_rows.append(i - 1)

        if i > limit:
            continue

        name_cell = ws.cell(row=i, column=1)
        fill = cell_fill_rgb(name_cell)
        bold = cell_is_bold(name_cell)
        label = row_label(ws, i, 1)

        if fill in GRAY_FILLS:
            gray_fill_rows.append(i - 1)
            if row_has_numbers(ws, i) or re.match(r"^итого", label, re.I):
                likely_subtotal_rows.append(i - 1)

        if fill in HIERARCHY_FILLS:
            hierarchy_fill_rows.append(i - 1)

        if i <= 20 and bold and (fill in HEADER_FILL_HINTS or fill in GRAY_FILLS or label):
            likely_header_rows.append(i - 1)

        if re.match(r"^итого", label, re.I) and (i - 1) not in likely_subtotal_rows:
            likely_subtotal_rows.append(i - 1)

    return {
        "likely_subtotal_rows": sorted(set(likely_subtotal_rows)),
        "likely_header_rows": sorted(set(likely_header_rows)),
        "gray_fill_rows": sorted(set(gray_fill_rows)),
        "hierarchy_fill_rows": sorted(set(hierarchy_fill_rows)),
        "hidden_rows": sorted(set(hidden_rows)),
    }


def probe_workbook(path: Path, sheet_name: str | None = None) -> dict:
    wb = load_workbook(path, data_only=True, read_only=False)
    used_sheet = pick_sheet(wb, sheet_name or None)
    if not used_sheet:
        return {"ok": False, "error": "no sheets"}

    ws = wb[used_sheet]
    row_count = ws.max_row or 0
    large_file = row_count > 5000
    style_scan_limit = STYLE_SCAN_LIMIT if large_file else row_count

    row_outline_levels: list[int] = []
    row_meta: list[dict] = []
    has_outline = False

    for i in range(1, row_count + 1):
        rd = ws.row_dimensions.get(i)
        outline = int(getattr(rd, "outline_level", 0) or 0) if rd else 0
        hidden = bool(getattr(rd, "hidden", False)) if rd else False
        if outline > 0:
            has_outline = True
        row_outline_levels.append(outline)

        if not large_file or outline > 0 or hidden or i <= style_scan_limit:
            name_cell = ws.cell(row=i, column=1)
            fill = cell_fill_rgb(name_cell)
            bold = cell_is_bold(name_cell)
            if outline > 0 or hidden or fill or bold:
                meta = {"i": i - 1, "outline_level": outline, "hidden": hidden, "bold": bold}
                if fill:
                    meta["fill_rgb"] = fill
                row_meta.append(meta)

    merged_ranges: list[str] = []
    for merged in getattr(ws, "merged_cells", None).ranges if getattr(ws, "merged_cells", None) else []:
        merged_ranges.append(str(merged))

    style_hints = build_style_hints(ws, row_count, style_scan_limit)
    sheet_names = list(wb.sheetnames)
    wb.close()

    if large_file:
        style_hints["large_file_mode"] = True
        style_hints["style_scan_rows"] = style_scan_limit

    return {
        "ok": True,
        "sheet_name": used_sheet,
        "sheet_names": sheet_names,
        "row_count": row_count,
        "large_file_mode": large_file,
        "row_outline_levels": row_outline_levels,
        "has_outline": has_outline,
        "merged_ranges": merged_ranges,
        "row_meta": row_meta,
        "style_hints": style_hints,
        "probe_engine": "openpyxl",
    }


def main() -> int:
    if len(sys.argv) < 2:
        print(json.dumps({"ok": False, "error": "usage: excel_probe.py <file.xlsx> [sheet]"}, ensure_ascii=False))
        return 1

    path = Path(sys.argv[1])
    sheet = sys.argv[2] if len(sys.argv) > 2 and sys.argv[2] else None

    if not path.exists():
        print(json.dumps({"ok": False, "error": f"file not found: {path}"}, ensure_ascii=False))
        return 1

    ext = path.suffix.lower()
    if ext not in {".xlsx", ".xlsm"}:
        print(json.dumps({"ok": False, "error": f"unsupported extension: {ext}"}, ensure_ascii=False))
        return 1

    try:
        result = probe_workbook(path, sheet)
        print(json.dumps(result, ensure_ascii=False))
        return 0 if result.get("ok") else 1
    except Exception as exc:
        print(json.dumps({"ok": False, "error": str(exc)}, ensure_ascii=False))
        return 1


if __name__ == "__main__":
    raise SystemExit(main())

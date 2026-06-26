# -*- coding: utf-8 -*-
"""Хитрые OS01 фикстуры: merge, hidden rows, gray subtotals, NBSP."""
from pathlib import Path
import subprocess
import sys

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font
except ImportError:
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
    import openpyxl
    from openpyxl.styles import PatternFill, Font

ROOT = Path(__file__).resolve().parents[2]
TRICKY = ROOT / 'tricky' / 'os_01'

GRAY = PatternFill('solid', fgColor='D9D9D9')
BOLD = Font(bold=True)


def _header_rows():
    return [
        ['Ведомость амортизации основных средств', '', '', '', '', '', '', ''],
        ['ООО Тест Fixture', '', '', '', '', '', '', ''],
        ['', '', '', '', '', '', '', ''],
        ['', 'На начало периода', '', 'За период', '', 'На конец периода', '', ''],
        ['', 'стоимость', 'амортизация', 'стоимость', 'амортизация', 'стоимость', 'амортизация', ''],
    ]


def _tree_rows():
    return [
        ['Здания', '', '', '', '', '', '', ''],
        ['РТК Волгоград', '', '', '', '', '', '', ''],
        ['ОП АБГ-Волгоград', '', '', '', '', '', '', ''],
        ['80-000001 Склад производственный инв. №12345 от 01.01.2020', 1000, 200, 50, 30, 1020, 230, ''],
        ['Машины', '', '', '', '', '', '', ''],
        ['РТК Москва', '', '', '', '', '', '', ''],
        ['ОП Центральный', '', '', '', '', '', '', ''],
        ['80-000002 Станок токарный инв. №67890 от 15.06.2019', 5000, 1200, 100, 80, 5020, 1280, ''],
    ]


def _write(rows, filename, sheet_name='Исходная выгрузка 01', setup=None):
    TRICKY.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for r_idx, row in enumerate(rows, 1):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    if setup:
        setup(ws)
    out = TRICKY / filename
    wb.save(out)
    print(f'Wrote {out}')


def gen_merged_title():
    title = [['ООО Тест — Объединённая шапка организации (merged A1:H3)', '', '', '', '', '', '', '']]
    padding = [['', '', '', '', '', '', '', ''] for _ in range(3)]
    rows = title + padding + _header_rows()[2:] + _tree_rows()

    def setup(ws):
        ws.merge_cells('A1:H3')
        ws['A1'].font = BOLD

    _write(rows, 'os01_merged_title.xlsx', setup=setup)


def gen_hidden_rows():
    rows = _header_rows() + _tree_rows()

    def setup(ws):
        # detail rows hidden (outline simulation via hidden flag)
        ws.row_dimensions[9].hidden = True
        ws.row_dimensions[10].hidden = True

    _write(rows, 'os01_hidden_rows.xlsx', setup=setup)


def gen_gray_subtotals():
    rows = _header_rows() + _tree_rows()
    rows.insert(6, ['Итого по группе Здания', 1000, 200, 50, 30, 1020, 230, ''])

    def setup(ws):
        for col in range(1, 9):
            ws.cell(row=7, column=col).fill = GRAY
            ws.cell(row=7, column=col).font = BOLD

    _write(rows, 'os01_gray_subtotals.xlsx', setup=setup)


def gen_numbers_nbsp():
    nbsp = '\u00a0'
    rows = _header_rows() + [
        ['Здания', '', '', '', '', '', '', ''],
        ['РТК Центр', '', '', '', '', '', '', ''],
        ['ОП Юг', '', '', '', '', '', '', ''],
        [
            f'80-000005 Объект с NBSP в числах инв. №55555 от 01.01.2020',
            f'1{nbsp}000,50',
            f'200,00',
            f'50,00',
            f'30,00',
            f'1{nbsp}020,50',
            f'230,00',
            '',
        ],
    ]
    _write(rows, 'numbers_nbsp.xlsx', setup=None)


def main():
    gen_merged_title()
    gen_hidden_rows()
    gen_gray_subtotals()
    gen_numbers_nbsp()


if __name__ == '__main__':
    main()
